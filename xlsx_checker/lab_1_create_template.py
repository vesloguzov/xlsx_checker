# -*- coding: UTF-8 -*-
import sys
import random
import datetime
import time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import *
from labs_utils import randomDate

reload(sys)
sys.setdefaultencoding('utf8')

def set_border_and_fill(ws, cell_range, fill=None):
    border_side = Side(border_style='thin', color='000000')
    border = Border(left=border_side,
                right=border_side,
                top=border_side,
                bottom=border_side)

    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if fill:
                cell.fill = fill
    return ws


def set_table_header(ws):
    headers_names = ["№", "Ф.И.О.", "Должность", "Дата поступления", "Оклад, руб", "Премия", "Подоходный налог", "Сумма к выдаче, руб", "Сумма к выдаче, $"]
    ws = set_border_and_fill(ws, 'A4:I4', fill=PatternFill("solid", fgColor="DDDDDD"))
    for i in range(1, 10):
        ws.cell(row=4, column=i).value = headers_names[i-1]

    return ws

def lab_1_create_template(ws, employees, positions):
    ws = set_table_header(ws)
    dollar_rate = 48
    shuffle_employees = list(employees)
    random.shuffle(shuffle_employees)

    positions = ["Директор", "Менеджер", "Бухгалтер", "Зам. директора", "Секетарь", "Водитель", "Строитель", "Секетарь", "Водитель", "Строитель"]

    ws.column_dimensions["A"].width = 5.0
    ws.column_dimensions["B"].width = 28.0
    ws.column_dimensions["C"].width = 16.0
    ws.column_dimensions["D"].width = 18.0
    ws.column_dimensions["E"].width = 18.0
    ws.column_dimensions["F"].width = 18.0
    ws.column_dimensions["G"].width = 20.0
    ws.column_dimensions["H"].width = 20.0
    ws.column_dimensions["I"].width = 20.0

    ws['B2'] = 'Расчет заработной платы сотрудников предприятия ООО "Изумруд"'
    ws.merge_cells('B2:I2')
    ws['B2'].alignment = Alignment(horizontal="center", vertical="center")

    for i in range(1, len(employees)+1):
        for j in range(1, 5):
            pos_i = i + 4
            ws.cell(row=pos_i, column=j).alignment = Alignment(horizontal="center", vertical="center")
            if j==1:
                ws.cell(row=pos_i, column=j).value = i
            if j == 2:
                ws.cell(row=pos_i, column=j).value = shuffle_employees[i-1]
            if j == 3:
                ws.cell(row=pos_i, column=j).value = positions[i-1]
            if j == 4:
                ws.cell(row=pos_i, column=j).value = randomDate()
                ws.cell(row=pos_i, column=j).number_format = 'DD/MM/YY'
                # ws.cell(row=pos_i, column=j).number_format = 'YYYY.MM.DD'

    set_border_and_fill(ws, str(ws.cell(row=len(employees)+7, column=2).coordinate)+":"+str(ws.cell(row=len(employees)+10, column=3).coordinate), fill=PatternFill("solid", fgColor="DDDDDD") )
    ws.cell(row=len(employees) + 7, column=2).value = "Курс доллара: "
    ws.cell(row=len(employees) + 7, column=3).value = dollar_rate

    ws.cell(row=len(employees) + 8, column=2).value = "Средняя зарплата, руб:"

    ws.cell(row=len(employees) + 9, column=2).value = "Максимальная зарплата, руб:"
    ws.cell(row=len(employees) + 10, column=2).value = "Минимальная зарплата, руб:"



    ws = set_border_and_fill(ws, 'A4:I11')
    ws = set_border_and_fill(ws, 'G12:I12')

    return ws