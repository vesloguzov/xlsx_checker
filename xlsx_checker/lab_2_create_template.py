# -*- coding: UTF-8 -*-
import sys
import random
import datetime
import time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import *
from openpyxl.drawing.image import Image

reload(sys)
sys.setdefaultencoding('utf8')

def set_border_and_fill(ws, cell_range, fill=None):
    border_side = Side(border_style='thin', color='000000')
    border = Border(left=border_side,
                right=border_side,
                top=border_side,
                bottom=border_side)

    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if fill:
                cell.fill = fill
    return ws

def create_table_ws_1(ws):
    ws.column_dimensions["A"].width = 17
    ws.column_dimensions["B"].width = 17
    ws['A1'] = 'График функции y=sin(x)'
    ws.merge_cells('A1:B2')
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A3'].value = 'x'
    ws['B3'].value = 'y=sin(x)'
    ws = set_border_and_fill(ws, 'A3:B3', fill=PatternFill("solid", fgColor="DDDDDD"))
    ws = set_border_and_fill(ws, 'A4:B28', None)

def create_table_ws_2(ws):
    ws.column_dimensions["A"].width = 17
    ws.column_dimensions["B"].width = 17
    ws['A1'] = 'График функции '
    ws.merge_cells('A1:B2')
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A3'].value = 'x'
    ws['B3'].value = 'y'
    ws = set_border_and_fill(ws, 'A3:B3', fill=PatternFill("solid", fgColor="DDDDDD"))
    ws = set_border_and_fill(ws, 'A4:B34', None)

    img = Image('img/lab_2_equation.png')
    ws.add_image(img, 'C1')
    ws.row_dimensions[1].height = 22.5
    ws.row_dimensions[2].height = 22.5

def lab_2_create_template(wb):
    ws1 = wb.active
    ws1.title = "Лист 1".decode('utf8')
    ws2 = wb.create_sheet(title="Лист 2".decode('utf8'))
    create_table_ws_1(ws1)
    create_table_ws_2(ws2)
    return wb