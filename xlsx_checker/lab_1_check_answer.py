# -*- coding: UTF-8 -*-
import sys
import json


from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.reader import reader
from openpyxl.chart.layout import Layout, ManualLayout
from labs_utils import range_is_date_format, range_is_money_rub_format, range_is_money_dollar_format, formulas_is_equal, approx_equal
from zipfile import ZipFile, ZIP_DEFLATED
from lxml.etree import fromstring, tostring
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

reload(sys)
sys.setdefaultencoding('utf8')

response_message = {}
response_message["formulas"] = {}
response_message["formats"] = {}
response_message["functions"] = {}
response_message["conditional_formatting"] = {}
response_message["charts"] = {}
response_message["errors"] = []

def range_is_formula_format(ws, range):
    rows = ws[range]
    for row in rows:
        for cell in row:
            if str(cell.value)[0] != '=':
                return False
    return True

def range_values_array(ws, range):
    arr = []
    rows = ws[range]
    for row in rows:
        for cell in row:
            arr.append(cell.value)
    return arr

def range_values_array_numeric(ws, range):
    arr = []
    rows = ws[range]
    for row in rows:
        for cell in row:
            arr.append(round(float(cell.value), 1))
    return arr

def calculate_correct_values(ws, range, employees, dollar_rate):
    data = {}
    data["names"] = employees
    try:
        rows = ws[range]
        values = []
        for row in rows:
            for cell in row:
                values.append(float(cell.value))
        data["salary"] = {"values": values}

        premium = []
        total = []
        income_tax = []
        amount_granted = []
        amount_granted_dollar = []

        premium_formula = []
        total_formula = []
        income_tax_formula = []
        amount_granted_formula = []
        amount_granted_dollar_formula = []



        for index, v in enumerate(data["salary"]["values"]):

            premium_val = round(v * 0.2, 1)
            premium.append(premium_val)
            premium_formula.append('=0,2*E'+str(index+5))

            total_val = round(premium_val + v, 1)
            total.append(total_val)
            total_formula.append('=F'+str(index+5)+'+E'+str(index+5))

            income_tax_val = round(total_val*0.13, 1)
            income_tax.append(income_tax_val)
            income_tax_formula.append('=G'+str(index+5)+'*0,13')

            amount_granted_val = round((total_val - income_tax_val), 1)
            amount_granted.append(amount_granted_val)
            amount_granted_formula.append('=G'+str(index+5)+'-H'+str(index+5))

            amount_granted_dollar_val = round(amount_granted_val/dollar_rate, 1)
            amount_granted_dollar.append(amount_granted_dollar_val)
            amount_granted_dollar_formula.append('=I'+str(index+5)+'/$C$14')

        data["premium"] = {"values": premium, "formula": premium_formula}
        data["total"] = {"values": total, "formula": total_formula}
        data["total_sum"] = { "values": sum(total), "formula": '=SUM(G5:G11)'}
        data["income_tax"] = {"values": income_tax, "formula": income_tax_formula}
        data["income_tax_sum"] = { "values": sum(income_tax), "formula": '=SUM(H5:H11)'}
        data["amount_granted"] = {"values": amount_granted, "formula": amount_granted_formula}
        data["amount_granted_sum"] = {"values": sum(amount_granted), "formula": '=SUM(I5:I11)'}
        data["amount_granted_dollar"] = {"values": amount_granted_dollar, "formula": amount_granted_dollar_formula}
        data["amount_granted_dollar_sum"] = {"values": sum(amount_granted_dollar), "formula": '=SUM(J5:J11)'}

        data["avg_salary"] = {"values": round(reduce(lambda x, y: x + y, amount_granted) / len(amount_granted), 2), "formula": '=AVERAGE(I5:I11)'}
        data["min_salary"] = {"values": round(min(amount_granted), 2), "formula": '=MIN(I5:I11)'}
        data["max_salary"] = {"values": round(max(amount_granted), 2), "formula": '=MAX(I5:I11)'}
        return data
    except:
        return False

def formulas_arrays_is_equal(f1, f2):
    clean_f1 = []
    clean_f2 = []
    for f in f1:
        clean_f1.append(f.replace(" ", "").lower().replace(".", ","))
    for f in f2:
        clean_f2.append(f.replace(" ", "").lower().replace(".", ","))

    return clean_f1 == clean_f2

def get_range_data(ws, range, ws_data_only):
    try:
        data = {}
        formula =  range_values_array(ws, range)
        data["formula"] = formula
        values = range_values_array_numeric(ws_data_only, range)
        data["values"]= values
        return data
    except:
        return False

def check_names(correct_data, student_data):
    if correct_data["names"] == student_data["names"]:
        return {'status': True, 'message': 'Names valid'}
    else:
        return {'status': False, 'message': 'Names invalid'}

def check_formats(student_ws):
    # Проверяем правильность форматирования дат поступления
    response_message["formats"]["dates"] = {}
    if range_is_date_format(student_ws, 'D5:D11')['status']:
        response_message["formats"]["dates"]["status"] = True
        response_message["formats"]["dates"]["message"] = "Форматирование дат поступления верно"
    else:
        response_message["formats"]["dates"]["status"] = False
        response_message["formats"]["dates"]["message"] = "Форматирование дат поступления неверно"

    # Проверяем правильность форматирования оклада
    response_message["formats"]["salary"] = {}
    if range_is_money_rub_format(student_ws, 'E5:E11')['status']:
        response_message["formats"]["salary"]["status"] = True
        response_message["formats"]["salary"]["message"] = "Форматирование оклада верно"
    else:
        response_message["formats"]["salary"]["status"] = False
        response_message["formats"]["salary"]["message"] = "Форматирование оклада неверно"

    # Проверяем правильность форматирования премии
    response_message["formats"]["premium"] = {}
    if range_is_money_rub_format(student_ws, 'F5:F11')['status']:
        response_message["formats"]["premium"]["status"] = True
        response_message["formats"]["premium"]["message"] = "Форматирование премии верно"
    else:
        response_message["formats"]["premium"]["status"] = False
        response_message["formats"]["premium"]["message"] = "Форматирование премии неверно"

    # Проверяем правильность форматирования итого
    response_message["formats"]["total"] = {}
    if range_is_money_rub_format(student_ws, 'G5:G11')['status']:
        response_message["formats"]["total"]["status"] = True
        response_message["formats"]["total"]["message"] = "Форматирование столбца Итого верно"
    else:
        response_message["formats"]["total"]["status"] = False
        response_message["formats"]["total"]["message"] = "Форматирование столбца Итого неверно"

    # Проверяем правильность форматирования подоходного налога
    response_message["formats"]["income_tax"] = {}
    if range_is_money_rub_format(student_ws, 'H5:H11')['status']:
        response_message["formats"]["income_tax"]["status"] = True
        response_message["formats"]["income_tax"]["message"] = "Форматирование подоходного налога верно"
    else:
        response_message["formats"]["income_tax"]["status"] = False
        response_message["formats"]["income_tax"]["message"] = "Форматирование подоходного налога неверно"

    # Проверяем правильность форматирования суммы к выдаче
    response_message["formats"]["amount_granted"] = {}
    if range_is_money_rub_format(student_ws, 'I5:I11')['status']:
        response_message["formats"]["amount_granted"]["status"] = True
        response_message["formats"]["amount_granted"]["message"] = "Форматирование суммы к выдаче верно"
    else:
        response_message["formats"]["amount_granted"]["status"] = False
        response_message["formats"]["amount_granted"]["message"] = "Форматирование суммы к выдаче неверно"

    # Проверяем правильность форматирования суммы к выдаче в долларах
    response_message["formats"]["amount_granted_dollar"] = {}
    if range_is_money_dollar_format(student_ws, 'J5:J11')['status']:
        response_message["formats"]["amount_granted_dollar"]["status"] = True
        response_message["formats"]["amount_granted_dollar"]["message"] = "Форматирование суммы к выдаче в долларах верно"
    else:
        response_message["formats"]["amount_granted_dollar"]["status"] = False
        response_message["formats"]["amount_granted_dollar"]["message"] = "Форматирование суммы к выдаче в долларах неверно"

def formula_contain_percent(formulas, percent):
    for formula in formulas:
        if not '0,'+str(percent) in formula and '%' in formula:
            return False
    return True

def is_formula_contain_percent(formulas, percent):
    if percent % 10 == 0:
        percent = str(percent)[0]
    print percent
    for formula in formulas:
        if formula.replace(" ", "")[0] != '=':
            if not '0,'+str(percent) in formula:
                if not '%' in formula:
                    return False
            return False

    return True

def check_formulas(ws, ws_read_only, correct_data):

    # Проверяем Итого
    total = get_range_data(ws, 'G5:G11', ws_read_only)
    response_message["formulas"]["total"] = {}
    if total and formulas_arrays_is_equal(total["formula"], correct_data["total"]["formula"]) and total["values"] == correct_data["total"]["values"]:
        response_message["formulas"]["total"]["status"] = True
        response_message["formulas"]["total"]["message"] = "Столбец итого посчитан верно"
    else:
        response_message["formulas"]["total"]["status"] = False
        response_message["formulas"]["total"]["message"] = "Столбец итого посчитан неверно"

    # Проверяем сумму к выдаче
    amount_granted = get_range_data(ws, 'I5:I11', ws_read_only)
    response_message["formulas"]["amount_granted"] = {}
    if amount_granted and formulas_arrays_is_equal(amount_granted["formula"], correct_data["amount_granted"]["formula"]) and amount_granted["values"] == correct_data["amount_granted"]["values"]:
        response_message["formulas"]["amount_granted"]["status"] = True
        response_message["formulas"]["amount_granted"]["message"] = "Столбец сумма к выдаче посчитан верно"
    else:
        response_message["formulas"]["amount_granted"]["status"] = False
        response_message["formulas"]["amount_granted"]["message"] = "Столбец сумма к выдаче посчитан неверно"

    # Проверяем сумму к выдаче в долларах
    amount_granted_dollar = get_range_data(ws, 'J5:J11', ws_read_only)
    response_message["formulas"]["amount_granted_dollar"] = {}
    if amount_granted_dollar and formulas_arrays_is_equal(amount_granted_dollar["formula"], correct_data["amount_granted_dollar"]["formula"]) and amount_granted_dollar["values"] == correct_data["amount_granted_dollar"]["values"]:
        response_message["formulas"]["amount_granted_dollar"]["status"] = True
        response_message["formulas"]["amount_granted_dollar"]["message"] = "Столбец сумма к выдаче в долларах посчитан верно"
    else:
        response_message["formulas"]["amount_granted_dollar"]["status"] = False
        response_message["formulas"]["amount_granted_dollar"]["message"] = "Столбец сумма к выдаче в долларах посчитан неверно"

    # Проверяем премии
    premium = get_range_data(ws, 'F5:F11', ws_read_only)

    response_message["formulas"]["premium"] = {}
    if premium and is_formula_contain_percent(premium["formula"], 20) and premium["values"] == correct_data["premium"]["values"]:
        response_message["formulas"]["premium"]["status"] = True
        response_message["formulas"]["premium"]["message"] = "Столбец премии посчитан верно"
    else:
        response_message["formulas"]["premium"]["status"] = False
        response_message["formulas"]["premium"]["message"] = "Столбец премии посчитан неверно"

    # Проверяем подоходный налог
    income_tax = get_range_data(ws, 'H5:H11', ws_read_only)
    response_message["formulas"]["income_tax"] = {}
    if income_tax and is_formula_contain_percent(income_tax["formula"], 13) and income_tax["values"] == correct_data["income_tax"]["values"]:
        response_message["formulas"]["income_tax"]["status"] = True
        response_message["formulas"]["income_tax"]["message"] = "Столбец подоходный налог посчитан верно"
    else:
        response_message["formulas"]["income_tax"]["status"] = False
        response_message["formulas"]["income_tax"]["message"] = "Столбец подоходный налог посчитан неверно"

def check_functions(ws, ws_read_only, correct_data):

    # Проверяем сумму Итого
    total_sum_cell = 'G12'
    response_message["functions"]["total_sum"] = {}
    if formulas_is_equal(ws[total_sum_cell].value, correct_data["total_sum"]["formula"]) and approx_equal(round(ws_read_only[total_sum_cell].value, 1), round(correct_data["total_sum"]["values"], 1)):
        response_message["functions"]["total_sum"]["status"] = True
        response_message["functions"]["total_sum"]["message"] = "Сумма по столбцу Итого посчитана верно"
    else:
        response_message["functions"]["total_sum"]["status"] = False
        response_message["functions"]["total_sum"]["message"] = "Сумма по столбцу Итого посчитана неверно"

    # Проверяем сумму подоходных налогов
    income_tax_sum_cell = 'H12'
    response_message["functions"]["income_tax_sum"] = {}
    if formulas_is_equal(ws[income_tax_sum_cell].value, correct_data["income_tax_sum"]["formula"]) and approx_equal(round(ws_read_only[income_tax_sum_cell].value, 1), round(correct_data["income_tax_sum"]["values"], 1)):
        response_message["functions"]["income_tax_sum"]["status"] = True
        response_message["functions"]["income_tax_sum"]["message"] = "Сумма подоходного налога посчитана верно"
    else:
        response_message["functions"]["income_tax_sum"]["status"] = False
        response_message["functions"]["income_tax_sum"]["message"] = "Сумма подоходного налога посчитана неверно"

    # Проверяем сумму сумм к выдаче
    amount_granted_sum_cell = 'I12'
    response_message["functions"]["amount_granted_sum"] = {}


    if formulas_is_equal(ws[amount_granted_sum_cell].value, correct_data["amount_granted_sum"]["formula"]) and approx_equal(round(ws_read_only[amount_granted_sum_cell].value, 1), round(correct_data["amount_granted_sum"]["values"], 1)):
        response_message["functions"]["amount_granted_sum"]["status"] = True
        response_message["functions"]["amount_granted_sum"]["message"] = "Сумма зарплат к выдаче в рублях посчитана верно"
    else:
        response_message["functions"]["amount_granted_sum"]["status"] = False
        response_message["functions"]["amount_granted_sum"]["message"] = "Сумма зарплат к выдаче в рублях посчитана неверно"


    # Проверяем сумму сумм к выдаче в долларах
    amount_granted_dollar_sum_cell = 'J12'
    response_message["functions"]["amount_granted_dollar_sum"] = {}
    
    if formulas_is_equal(ws[amount_granted_dollar_sum_cell].value, correct_data["amount_granted_dollar_sum"]["formula"]) and approx_equal(int(ws_read_only[amount_granted_dollar_sum_cell].value), int(correct_data["amount_granted_dollar_sum"]["values"])):
        response_message["functions"]["amount_granted_dollar_sum"]["status"] = True
        response_message["functions"]["amount_granted_dollar_sum"]["message"] = "Сумма зарплат к выдаче в долларах посчитана верно"
    else:
        response_message["functions"]["amount_granted_dollar_sum"]["status"] = False
        response_message["functions"]["amount_granted_dollar_sum"]["message"] = "Сумма зарплат к выдаче в долларах посчитана неверно"

    # Проверяем среднее значение суммы к выдаче
    avg_salary_sum_cell = 'C15'
    response_message["functions"]["avg_salary"] = {}
    if formulas_is_equal(ws[avg_salary_sum_cell].value, correct_data["avg_salary"]["formula"]) and approx_equal(round(ws_read_only[avg_salary_sum_cell].value, 1), round(correct_data["avg_salary"]["values"], 1)):
        response_message["functions"]["avg_salary"]["status"] = True
        response_message["functions"]["avg_salary"]["message"] = "Среднее значение зарплаты посчитано верно"
    else:
        response_message["functions"]["avg_salary"]["status"] = False
        response_message["functions"]["avg_salary"]["message"] = "Среднее значение зарплаты посчитано неверно"

    # Проверяем максимальное значение суммы к выдаче
    max_salary_sum_cell = 'C16'
    response_message["functions"]["max_salary"] = {}
    if formulas_is_equal(ws[max_salary_sum_cell].value, correct_data["max_salary"]["formula"]) and approx_equal(round(ws_read_only[max_salary_sum_cell].value, 1), round(correct_data["max_salary"]["values"], 1)):
        response_message["functions"]["max_salary"]["status"] = True
        response_message["functions"]["max_salary"]["message"] = "Максимальное значение зарплаты посчитано верно"
    else:
        response_message["functions"]["max_salary"]["status"] = False
        response_message["functions"]["max_salary"]["message"] = "Максимальное значение зарплаты посчитано неверно"

    # Проверяем минимальное значение суммы к выдаче
    min_salary_sum_cell = 'C17'
    response_message["functions"]["min_salary"] = {}
    if formulas_is_equal(ws[min_salary_sum_cell].value, correct_data["min_salary"]["formula"]) and approx_equal(round(ws_read_only[min_salary_sum_cell].value, 1), round(correct_data["min_salary"]["values"], 1)):
        response_message["functions"]["min_salary"]["status"] = True
        response_message["functions"]["min_salary"]["message"] = "Минимальное значение зарплаты посчитано верно"
    else:
        response_message["functions"]["min_salary"]["status"] = False
        response_message["functions"]["min_salary"]["message"] = "Минимальное значение зарплаты посчитано неверно"

def check_ws_have_rule(ws, cells_range, operator, formula_value):
    for rule in ws.conditional_formatting.cf_rules.items():
        if cells_range == rule[0]:
            if operator == rule[1][0].operator:
                if int(rule[1][0].formula[0]) == formula_value:
                    return True
    return False

def check_bar_graphic(filename, data_x, data_y):
    analyze = {}
    analyze["bar_chart"] = {}
    analyze["bar_chart"]["errors"] = []
    analyze["bar_chart"]["data_x"] = {}
    analyze["bar_chart"]["data_y"] = {}
    analyze["bar_chart"]["title_x"] = {}
    analyze["bar_chart"]["title_y"] = {}
    analyze["bar_chart"]["chart_title"] = {}

    sourceFile = ZipFile(filename, 'r')
    charts = []; [charts.append(sourceFile.read(ch)) for ch in sourceFile.namelist() if 'charts/chart' in ch]
    charts_objects = []
    for chart in charts:
        try:
            clean_chart = fromstring(chart)
            for bad in clean_chart.xpath(".//c:extLst", namespaces={'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart'}):
                bad.getparent().remove(bad)
            clean_chart = tostring(clean_chart)
            l = reader(clean_chart)

            charts_objects.append(l)
        except:
            analyze["bar_chart"]["errors"].append('График типа "Гистограмма" не обнаружен!')
            return analyze
    obj = {}
    for chart in charts_objects:
        if chart.tagname == 'barChart':
            obj = chart
            break

    if obj == {}:
        analyze["bar_chart"]["errors"].append('График типа "Гистограмма" не обнаружен!')
    else:
        if obj.title != None:
            analyze["bar_chart"]["chart_title"]["message"] = 'Имя графика присвоено'
            analyze["bar_chart"]["chart_title"]["status"] = True
        else:
            analyze["bar_chart"]["chart_title"]["message"] = 'Имя графика не присвоено'
            analyze["bar_chart"]["chart_title"]["status"] = False

        try:
            for s in obj.ser:
                if data_x in s.cat.strRef.f.replace(" ", ""):
                    analyze["bar_chart"]["data_x"]["message"] = 'Данные для оси x выбраны верно'
                    analyze["bar_chart"]["data_x"]["status"] = True
                else:
                    analyze["bar_chart"]["data_x"]["message"] = 'Данные для оси x выбраны неверно'
                    analyze["bar_chart"]["data_x"]["status"] = False

        except:
            analyze["bar_chart"]["data_x"]["message"] = 'Данные для оси x выбраны неверно'
            analyze["bar_chart"]["data_x"]["status"] = False
        try:
            for s in obj.ser:
                if data_y in s.val.numRef.f.replace(" ", ""):
                    analyze["bar_chart"]["data_y"]["message"] = 'Данные для оси y выбраны верно'
                    analyze["bar_chart"]["data_y"]["status"] = True
                else:
                    analyze["bar_chart"]["data_y"]["message"] = 'Данные для оси y выбраны неверно'
                    analyze["bar_chart"]["data_y"]["status"] = False
        except:
            analyze["bar_chart"]["data_y"]["message"] = 'Данные для оси y выбраны неверно'
            analyze["bar_chart"]["data_y"]["status"] = False

        try:
            if obj.x_axis.title != None:
                analyze["bar_chart"]["title_x"]["message"] = 'Подпись оси x выполнена'
                analyze["bar_chart"]["title_x"]["status"] = True
            else:
                analyze["bar_chart"]["title_x"]["message"] = 'Подпись оси x не выполнена'
                analyze["bar_chart"]["title_x"]["status"] = False
        except:
            analyze["bar_chart"]["title_x"]["message"] = 'Подпись оси x не выполнена'
            analyze["bar_chart"]["title_x"]["status"] = False

        try:
            if obj.y_axis.title != None:
                analyze["bar_chart"]["title_y"]["message"] = 'Подпись оси y выполнена'
                analyze["bar_chart"]["title_y"]["status"] = True
            else:
                analyze["bar_chart"]["title_y"]["message"] = 'Подпись оси y не выполнена'
                analyze["bar_chart"]["title_y"]["status"] = False
        except:
            analyze["bar_chart"]["title_y"]["message"] = 'Подпись оси y не выполнена'
            analyze["bar_chart"]["title_y"]["status"] = False

    return analyze

def lab_1_check_answer(student_wb, student_wb_data_only, file):
    # response_message = {}
    # response_message["formulas"] = {}
    # response_message["formats"] = {}
    # response_message["functions"] = {}
    # response_message["conditional_formatting"] = {}
    # response_message["charts"] = {}
    response_message["errors"] = []

    dollar_rate = 48
    employees = ["Иванов И.М.", "Коробова П.Н", "Морозов И.Р.", "Петров Г.Т.", "Ромашова П.Т.", "Смирнов С.И.", "Соколова О.С."]

    student_ws = student_wb[student_wb.get_sheet_names()[0]]
    ws_read_only = student_wb_data_only[student_wb_data_only.get_sheet_names()[0]]
    correct_values_data = calculate_correct_values(student_ws, 'E5:E11', sorted(employees), dollar_rate)
    if correct_values_data:
    
        # Проверяем правильность ФИО (в т.ч. сортировку)
        check_formats(student_ws)

        # Проверем правильность формул и значений
        check_formulas(student_ws, ws_read_only, correct_values_data)

        # Проверяем правильность функций
        check_functions(student_ws, ws_read_only, correct_values_data)

        response_message["charts"] = check_bar_graphic(file, '$B$5:$B$11', '$I$5:$I$11')

        # Проверем правильность условного форматирования
        if check_ws_have_rule(student_ws, 'I5:I11', 'lessThan', 5500):
            response_message["conditional_formatting"]["message"] = 'Условное форматирование выполнено верно'
            response_message["conditional_formatting"]["status"] = True
        else:
            response_message["conditional_formatting"]["message"] = 'Условное форматирование выполнено неверно'
            response_message["conditional_formatting"]["status"] = False
    else:
        response_message["errors"].append('Неверно заполнен столбец "Оклад"')
    return response_message