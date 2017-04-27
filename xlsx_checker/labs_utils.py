# -*- coding: UTF-8 -*-
import sys
import random
import datetime
import time

from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.reader import reader
from openpyxl.chart.layout import Layout, ManualLayout

reload(sys)
sys.setdefaultencoding('utf8')

def range_is_date_format(ws, range):
    rows = ws[range]
    for row in rows:
        for cell in row:
            if not cell.is_date:
                return {'status': False, 'message': 'Dates invalid'}
    return {'status': True, 'message': 'Dates valid'}

def range_is_money_rub_format(ws, range):
    rows = ws[range]
    for row in rows:
        for cell in row:
            if not '₽' in cell.number_format:
                return {'status': False, 'message': 'Money rub format invalid'}
    return {'status': True, 'message': 'Money rub format valid'}

def range_is_money_dollar_format(ws, range):
    rows = ws[range]
    for row in rows:
        for cell in row:
            if not '$' in cell.number_format:
                return {'status': False, 'message': 'Money dollar format invalid'}
    return {'status': True, 'message': 'Money dollar format valid'}

def formulas_is_equal(f1, f2):
    if f1 and f2:
        f1 = f1.replace(" ", "").lower().replace(".", ",")
        f2 = f2.replace(" ", "").lower().replace(".", ",")

        return f1 == f2

    else: return False

def strTimeProp(start, end, format, prop):
    stime = time.mktime(time.strptime(start, format))
    etime = time.mktime(time.strptime(end, format))
    ptime = stime + prop * (etime - stime)
    return time.strftime(format, time.localtime(ptime))

startDate = "01.01.2000"
endDate = "01.01.2018"

def randomDate(start=startDate, end=endDate):
    prop = random.random()
    return strTimeProp(start, end, '%d.%m.%Y', prop)