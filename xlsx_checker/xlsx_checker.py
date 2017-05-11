# -*- coding: utf-8 -*-

import pkg_resources


import hashlib
import logging
import mimetypes
import os
import uuid
import sys
import json
import random

from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.reader import reader
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.styles import *
# from docx import Document

from xblock.core import XBlock
from xblock.fields import Scope, Integer, String, JSONField
from xblock.fragment import Fragment
from django.core.files import File
from django.core.files.storage import default_storage
from django.core.exceptions import PermissionDenied
from django.views.decorators.csrf import csrf_exempt

from student.models import user_by_anonymous_id
from submissions import api as submissions_api
from submissions.models import StudentItem as SubmissionsStudent

from functools import partial
from xmodule.util.duedate import get_extended_due_date
from webob.response import Response


from .utils import (
    load_resource,
    render_template,
    load_resources,
    )

from lab_1_create_template import lab_1_create_template
from lab_1_check_answer import lab_1_check_answer

from lab_2_create_template import lab_2_create_template
from lab_2_check_answer import lab_2_check_answer

from lab_3_create_template import lab_3_create_template
from lab_3_check_answer import lab_3_check_answer

log = logging.getLogger(__name__)
BLOCK_SIZE = 8 * 1024

reload(sys)
sys.setdefaultencoding('utf8')

class XlsxCheckerXBlock(XBlock):
    """
    TO-DO: document what your XBlock does.
    """
    lab_scenario = Integer(
        display_name=u"Номер сценария",
        help=(u"Номер сценария",
              u"Номер сценария"),
        default=9999,
        scope=Scope.settings
    )

    scenarios_settings = JSONField(
        display_name=u"Настройки сценария",
        help=u"Настройки сценария",
        default={
            "1": {
                  "instruction_name": "Лабораторная 1. Указания к работе.docx", 
                  "template_name": "lab1_template.xlsx",
                  "correct_name": "lab1_correct.xlsx",
                  "title": "Формулы, функции и диаграммы в процессоре Microsoft Office Excel"
                 },
            "2": {
                  "instruction_name": "Лабораторная 2. Указания к работе.docx", 
                  "template_name": "lab2_template.xlsx",
                  "correct_name": "lab2_correct.xlsx",
                  "title": "Построение графиков функций"
                 },
            "3": {
                  "instruction_name": "Лабораторная 3. Указания к работе.docx",
                  "template_name": "lab3_template.xlsx",
                  "correct_name": "lab3_correct.xlsx",
                  "title": "Сортировка, фильтры и промежуточные итоги"
                 },
        },
        scope=Scope.settings
    )

    instruction_link = String(
         default='', scope=Scope.settings,
         help='Link for instruction download',
        )

    template_link = String(
         default='', scope=Scope.settings,
         help='Link for template download',
        )

    correct_link = String(
         default='', scope=Scope.settings,
         help='Link for correct file',
        )

    xlsx_analyze = JSONField(
         default={}, 
         scope=Scope.user_state,
         help='Analyze document',
        )



    correct_xlsx_uid = String(
         default='', scope=Scope.settings,
         help='Correct file from teacher',
        )

    correct_xlsx_name = String(
         default='', scope=Scope.settings,
         help='Name of correct file from teacher',
        )

    student_xlsx_uid = String(
         default='', scope=Scope.user_state,
         help='Studen file from student',
        )
    student_xlsx_name = String(
         default='', scope=Scope.user_state,
         help='Name of student file from student',
        )

    display_name = String(
        display_name=u"Название",
        help=u"Название задания, которое увидят студенты.",
        default=u'Проверка MS Excel',
        scope=Scope.settings
    )

    question = String(
        # TODO: list
        display_name=u"Вопрос",
        help=u"Текст задания.",
        default=u"Лабораторные работы MS Excel",
        scope=Scope.settings
    )

    weight = Integer(
        display_name=u"Максимальное количество баллов",
        help=(u"Максимальное количество баллов",
              u"которое может получить студент."),
        default=10,
        scope=Scope.settings
    )

    #TODO: 1!
    max_attempts = Integer(
        display_name=u"Максимальное количество попыток",
        help=u"",
        default=10,
        scope=Scope.settings
    )
    
    attempts = Integer(
        display_name=u"Количество использованных попыток",
        help=u"",
        default=0,
        scope=Scope.user_state
    )

    points = Integer(
        display_name=u"Текущее количество баллов студента",
        default=None,
        scope=Scope.user_state
    )

    def resource_string(self, path):
        """Handy helper for getting resources from our kit."""
        data = pkg_resources.resource_string(__name__, path)
        return data.decode("utf8")

    # TO-DO: change this view to display your data your own way.
    def student_view(self, context=None):
        context = {
            "display_name": self.display_name,
            "weight": self.weight,
            "question": self.question,
            "student_xlsx_name": self.student_xlsx_name,
            "points": self.points,
            "attempts": self.attempts,
            "instruction_link": self.runtime.local_resource_url(self, 'public/instructions/' + self.instruction_link),
            "template_link": self.runtime.local_resource_url(self, 'public/templates/' + self.template_link),
            "lab_scenario": self.lab_scenario,
            "xlsx_analyze": self.xlsx_analyze,
            "download_template_icon": self.runtime.local_resource_url(self, 'public/images/download_template_icon.png'),
            "download_instruction_icon": self.runtime.local_resource_url(self, 'public/images/download_instruction_icon.png'),
        }

        if self.max_attempts != 0:
            context["max_attempts"] = self.max_attempts

        if self.past_due():
            context["past_due"] = True

        if answer_opportunity(self):
            context["answer_opportunity"] = True

        fragment = Fragment()
        fragment.add_content(
            render_template(
                "static/html/xlsx_checker.html",
                context
            )
        )

        js_urls = (
            "static/js/src/xlsx_checker.js",
            )

        css_urls = (
            "static/css/xlsx_checker.css",
            )

        load_resources(js_urls, css_urls, fragment)

        fragment.initialize_js('XlsxCheckerXBlock', {'xlsx_analyze': self.xlsx_analyze, 'lab_scenario': self.lab_scenario, 'student_xlsx_name': self.student_xlsx_name})
        return fragment

    def studio_view(self, context=None):

        scenarios = []
        for index,key in enumerate(self.scenarios_settings.keys()):
            element = {}
            element["title"] = self.scenarios_settings[str(index+1)]["title"]
            element["number"] = str(index+1)
            scenarios.append(element)

        context = {
            "display_name": self.display_name,
            "weight": self.weight,
            "question": self.question,
            "max_attempts": self.max_attempts,
            "lab_scenario": self.lab_scenario,
            "scenarios": scenarios,
        }

        fragment = Fragment()
        fragment.add_content(
            render_template(
                "static/html/xlsx_checker_studio.html",
                context
            )
        )

        js_urls = (
            "static/js/src/xlsx_checker_studio.js",
            )

        css_urls = (
            "static/css/xlsx_checker_studio.css",
            )

        load_resources(js_urls, css_urls, fragment)

        fragment.initialize_js('XlsxCheckerXBlock')
        return fragment


    # TO-DO: change this to create the scenarios you'd like to see in the
    # workbench while developing your XBlock.
    @staticmethod
    def workbench_scenarios():
        """A canned scenario for display in the workbench."""
        return [
            ("XlsxCheckerXBlock",
             """<xlsx_checker/>
             """),
            ("Multiple XlsxCheckerXBlock",
             """<vertical_demo>
                <xlsx_checker/>
                <xlsx_checker/>
                <xlsx_checker/>
                </vertical_demo>
             """),
        ]

    @XBlock.json_handler
    def student_submit(self, data, suffix=''):

        def check_answer():
            return random.randint(10, 100)

        result = {}
        student_path = self._students_storage_path(self.student_xlsx_uid, self.student_xlsx_name)
        
        self.xlsx_analyze["errors"] = []

        if str(self.lab_scenario) == "1":
            try:
                student_wb =  load_workbook(default_storage.open(student_path))
                student_wb_data_only =  load_workbook(default_storage.open(student_path), data_only=True)
                result = lab_1_check_answer(student_wb, student_wb_data_only, default_storage.open(student_path))
                self.xlsx_analyze = result
            except:
                self.xlsx_analyze["errors"].append("Ошибка открытия файла")

        if str(self.lab_scenario) == "2":
            try:
                correct_wb = load_workbook('/home/edx/edxwork/xlsx_checker/xlsx_checker/corrects/lab2_correct.xlsx')
                correct_wb_data_only =  load_workbook('/home/edx/edxwork/xlsx_checker/xlsx_checker/corrects/lab2_correct.xlsx', data_only=True)
                student_wb =  load_workbook(default_storage.open(student_path))
                student_wb_data_only =  load_workbook(default_storage.open(student_path), data_only=True)
                result = lab_2_check_answer(correct_wb, correct_wb_data_only, student_wb, student_wb_data_only, default_storage.open(student_path))
                self.xlsx_analyze = result
            except:
                self.xlsx_analyze["errors"].append("Ошибка открытия файла")

        if str(self.lab_scenario) == "3":
            try:
                correct_wb = load_workbook('/home/edx/edxwork/xlsx_checker/xlsx_checker/corrects/lab3_correct.xlsx')
                correct_wb_data_only =  load_workbook('/home/edx/edxwork/xlsx_checker/xlsx_checker/corrects/lab3_correct.xlsx', data_only=True)
                # Проверка
                student_wb =  load_workbook(default_storage.open(student_path))
                student_wb_data_only =  load_workbook(default_storage.open(student_path), data_only=True)
                result = lab_3_check_answer(correct_wb, correct_wb_data_only, student_wb, student_wb_data_only)
                self.xlsx_analyze = result
            except:
                self.xlsx_analyze["errors"].append("Ошибка открытия файла")

        grade_global = check_answer()
        self.points = grade_global
        self.points = grade_global * self.weight / 100
        self.points = int(round(self.points))
        self.attempts += 1
        self.runtime.publish(self, 'grade', {
            'value': self.points,
            'max_value': self.weight,
        })
        res = {"success_status": 'ok', "points": self.points, "weight": self.weight, "attempts": self.attempts, "max_attempts": self.max_attempts, "xlsx_analyze": self.xlsx_analyze }
        return res

    @XBlock.json_handler
    def studio_submit(self, data, suffix=''):
        self.display_name = data.get('display_name')
        self.question = data.get('question')
        self.weight = data.get('weight')
        self.max_attempts = data.get('max_attempts')
        self.lab_scenario = data.get('lab_scenario')

        self.instruction_link = self.scenarios_settings[str(self.lab_scenario)]["instruction_name"]
        self.template_link = self.scenarios_settings[str(self.lab_scenario)]["template_name"]
        self.correct_link = self.runtime.local_resource_url(self, 'corrects/' + self.scenarios_settings[str(self.lab_scenario)]["correct_name"])

        self.display_name = 'Проверка MS Excel. ' + self.scenarios_settings[str(self.lab_scenario)]["title"]

        if str(self.lab_scenario) == "1":
            template_wb = Workbook()
            template_ws = template_wb.active
            template_ws = lab_1_create_template(template_ws)
            template_wb.save('/home/edx/edxwork/xlsx_checker/xlsx_checker/public/templates/' + self.template_link)
            

        if str(self.lab_scenario) == "2":
            
            template_wb = Workbook()

            # Сделать нормально
            img = self.runtime.local_resource_url(self, 'public/img/lab_2_equation.png')
            img = '/home/edx/edxwork/xlsx_checker/xlsx_checker/public/images/lab_2_equation.png'

            template_wb = lab_2_create_template(template_wb, img)
            template_wb.save('/home/edx/edxwork/xlsx_checker/xlsx_checker/public/templates/' + self.template_link)

        if str(self.lab_scenario) == "3":
            template_wb = Workbook()
            template_ws = template_wb.active
            template_ws = lab_3_create_template(template_ws)
            template_wb.save('/home/edx/edxwork/xlsx_checker/xlsx_checker/public/templates/' + self.template_link)

        return {'result': 'success'}

    @XBlock.handler
    def student_filename(self, request, suffix=''):
        return Response(json_body={'student_filename': self.student_xlsx_name})

    @XBlock.handler
    def download_student_file(self, request, suffix=''):
        path = self._students_storage_path(self.student_xlsx_uid, self.student_xlsx_name)
        return self.download(
            path,
            mimetypes.guess_type(self.student_xlsx_name)[0],
            self.student_xlsx_name
        )


    @XBlock.handler
    def download_instruction(self, request, suffix=''):
        path = self.runtime.local_resource_url(self, 'public/instructions/' + self.instruction_link)
        return self.download(
            path,
            'docx',
            self.instruction_link
        )

    def is_course_staff(self):
        return getattr(self.xmodule_runtime, 'user_is_staff', False)

    @XBlock.handler
    def upload_student_file(self, request, suffix=''):
        upload = request.params['studentFile']
        self.student_xlsx_name = upload.file.name
        self.student_xlsx_uid = uuid.uuid4().hex
        path = self._students_storage_path(self.student_xlsx_uid, self.student_xlsx_name)
        if not default_storage.exists(path):
            default_storage.save(path, File(upload.file))
        obj = path
        return Response(json_body=obj)
    

    def _file_storage_path(self, uid, filename):
        # pylint: disable=no-member
        """
        Get file path of storage.
        """
        path = (
            '{loc.org}/{loc.course}/{loc.block_type}/templates'
            '/{uid}{ext}'.format(
                loc=self.location,
                uid= uid,
                ext=os.path.splitext(filename)[1]
            )
        )
        return path

    def _students_storage_path(self, uid, filename):
        path = (
            '{loc.org}/{loc.course}/{loc.block_type}/students'
            '/{uid}{ext}'.format(
                loc=self.location,
                uid=uid,
                ext=os.path.splitext(filename)[1]
            )
        )
        return path


    def download(self, path, mime_type, filename, require_staff=False):
        """
        Return a file from storage and return in a Response.
        """
        try:
            file_descriptor = default_storage.open(path)
            app_iter = iter(partial(file_descriptor.read, BLOCK_SIZE), '')
            return Response(
                app_iter=app_iter,
                content_type=mime_type,
                content_disposition="attachment; filename=" + filename.encode('utf-8'))
        except IOError:
            if require_staff:
                return Response(
                    "Sorry, assignment {} cannot be found at"
                    " {}. Please contact {}".format(
                        filename.encode('utf-8'), path, settings.TECH_SUPPORT_EMAIL
                    ),
                    status_code=404
                )
            return Response(
                "Sorry, the file you uploaded, {}, cannot be"
                " found. Please try uploading it again or contact"
                " course staff".format(filename.encode('utf-8')),
                status_code=404
            )

    def past_due(self):
            """
            Проверка, истекла ли дата для выполнения задания.
            """
            due = get_extended_due_date(self)
            if due is not None:
                if _now() > due:
                    return False
            return True

    def is_course_staff(self):
        """
        Проверка, является ли пользователь автором курса.
        """
        return getattr(self.xmodule_runtime, 'user_is_staff', False)

    def is_instructor(self):
        """
        Проверка, является ли пользователь инструктором.
        """
        return self.xmodule_runtime.get_user_role() == 'instructor'

def _now():
    """
    Получение текущих даты и времени.
    """
    return datetime.datetime.utcnow().replace(tzinfo=pytz.utc)

def answer_opportunity(self):
    """
    Возможность ответа (если количество сделанных попыток меньше заданного).
    """
    if self.max_attempts <= self.attempts and self.max_attempts != 0:
        return False
    else:
        return True

def require(assertion):
    """
    Raises PermissionDenied if assertion is not true.
    """
    if not assertion:
        raise PermissionDenied
